"""
Generate formatted Excel report from divergence detection results.
Reads JSON from output/divergence_results.json, outputs output/divergence_report.xlsx.

Usage:
    python scripts/generate_report.py
"""

import json
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ─── Styling ───────────────────────────────────────────────────────────────

HEADER_FONT = Font(bold=True, size=10, color="FFFFFF")
HEADER_FILL_BEAR = PatternFill("solid", fgColor="C0392B")
HEADER_FILL_BULL = PatternFill("solid", fgColor="27AE60")
SUCCESS_FILL = PatternFill("solid", fgColor="D5F5E3")
FAIL_FILL = PatternFill("solid", fgColor="FADBD8")
BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

BEARISH_HEADERS = [
    "#",
    "BTC 開始時間", "BTC 開始價格",
    "BTC HH 時間", "BTC HH 價格",
    "BTC 判斷失敗價格", "BTC 頸線價位",
    "BTC 頸線跌破時間",
    "BTC Rejection 時間", "BTC Rejection 價格",
    "ETH 開始時間", "ETH 開始價格",
    "ETH 高點時間", "ETH 高點價格",
    "ETH 判斷失敗價格", "ETH 頸線價位",
    "ETH 頸線跌破時間",
    "ETH Rejection 時間", "ETH Rejection 價格",
    "結果", "BTC 之後跌幅%", "ETH 之後跌幅%", "備註",
]

BULLISH_HEADERS = [
    "#",
    "BTC 開始時間", "BTC 開始價格",
    "BTC LL 時間", "BTC LL 價格",
    "BTC 判斷失敗價格", "BTC 頸線價位",
    "BTC 頸線突破時間",
    "BTC Rejection 時間", "BTC Rejection 價格",
    "ETH 開始時間", "ETH 開始價格",
    "ETH 低點時間", "ETH 低點價格",
    "ETH 判斷失敗價格", "ETH 頸線價位",
    "ETH 頸線突破時間",
    "ETH Rejection 時間", "ETH Rejection 價格",
    "結果", "BTC 之後漲幅%", "ETH 之後漲幅%", "備註",
]


def bearish_to_row(idx: int, case: dict) -> list:
    return [
        idx,
        case["btc_start_time"], case["btc_start_price"],
        case["btc_hh_time"], case["btc_hh_price"],
        case["btc_failure_price"], case["btc_neckline"],
        case["btc_neckline_break_time"],
        case["btc_rejection_time"], case["btc_rejection_price"],
        case["eth_start_time"], case["eth_start_price"],
        case["eth_high_time"], case["eth_high_price"],
        case["eth_failure_price"], case["eth_neckline"],
        case["eth_neckline_break_time"],
        case["eth_rejection_time"], case["eth_rejection_price"],
        case["result"], case["btc_drop_pct"], case["eth_drop_pct"], case["note"],
    ]


def bullish_to_row(idx: int, case: dict) -> list:
    return [
        idx,
        case["btc_start_time"], case["btc_start_price"],
        case["btc_ll_time"], case["btc_ll_price"],
        case["btc_failure_price"], case["btc_neckline"],
        case["btc_neckline_break_time"],
        case["btc_rejection_time"], case["btc_rejection_price"],
        case["eth_start_time"], case["eth_start_price"],
        case["eth_low_time"], case["eth_low_price"],
        case["eth_failure_price"], case["eth_neckline"],
        case["eth_neckline_break_time"],
        case["eth_rejection_time"], case["eth_rejection_price"],
        case["result"], case["btc_rise_pct"], case["eth_rise_pct"], case["note"],
    ]


def write_sheet(ws, headers: list, rows: list, header_fill: PatternFill):
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = header_fill
        cell.border = BORDER
        cell.alignment = CENTER

    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = BORDER
            cell.alignment = CENTER if col_idx <= 1 or col_idx >= len(headers) - 2 else LEFT

            if col_idx == len(headers) - 3:  # 結果 column
                if value == "成功":
                    cell.fill = SUCCESS_FILL
                elif value == "失敗":
                    cell.fill = FAIL_FILL
                cell.font = Font(bold=True)

    for col_idx in range(1, len(headers) + 1):
        max_len = len(str(headers[col_idx - 1]))
        for row_idx in range(2, len(rows) + 2):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 22)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    ws.freeze_panes = "B2"


def write_summary(ws, bearish: list, bullish: list, params: dict):
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18

    title_font = Font(bold=True, size=14)
    label_font = Font(bold=True, size=11)

    ws.cell(row=1, column=1, value="Cross-Asset Divergence Report").font = title_font

    ws.cell(row=3, column=1, value="Parameters").font = label_font
    row = 4
    for k, v in params.items():
        ws.cell(row=row, column=1, value=k)
        ws.cell(row=row, column=2, value=str(v))
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Summary").font = label_font
    row += 1

    bear_success = sum(1 for c in bearish if c["result"] == "成功")
    bear_fail = len(bearish) - bear_success
    bull_success = sum(1 for c in bullish if c["result"] == "成功")
    bull_fail = len(bullish) - bull_success

    headers = ["", "Bearish", "Bullish"]
    for ci, h in enumerate(headers):
        ws.cell(row=row, column=ci + 1, value=h).font = Font(bold=True)
    row += 1

    data_rows = [
        ["Total", len(bearish), len(bullish)],
        ["成功", bear_success, bull_success],
        ["失敗", bear_fail, bull_fail],
        ["勝率",
         f"{bear_success / len(bearish) * 100:.1f}%" if bearish else "N/A",
         f"{bull_success / len(bullish) * 100:.1f}%" if bullish else "N/A"],
    ]
    for dr in data_rows:
        for ci, v in enumerate(dr):
            ws.cell(row=row, column=ci + 1, value=v)
        row += 1


def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    output_dir = os.path.join(script_dir, "..", "output")
    results_path = os.path.join(output_dir, "divergence_results.json")

    if not os.path.exists(results_path):
        print("Error: divergence_results.json not found. Run detect_divergence.py first.")
        return

    with open(results_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    bearish = data.get("bearish", [])
    bullish = data.get("bullish", [])
    params = data.get("params", {})

    wb = Workbook()

    ws_summary = wb.active
    ws_summary.title = "Summary"
    write_summary(ws_summary, bearish, bullish, params)

    ws_bear = wb.create_sheet("Bearish Divergence")
    bear_rows = [bearish_to_row(i + 1, c) for i, c in enumerate(bearish)]
    write_sheet(ws_bear, BEARISH_HEADERS, bear_rows, HEADER_FILL_BEAR)

    ws_bull = wb.create_sheet("Bullish Divergence")
    bull_rows = [bullish_to_row(i + 1, c) for i, c in enumerate(bullish)]
    write_sheet(ws_bull, BULLISH_HEADERS, bull_rows, HEADER_FILL_BULL)

    report_path = os.path.join(output_dir, "divergence_report.xlsx")
    wb.save(report_path)
    print(f"Report saved -> {report_path}")
    print(f"  Bearish: {len(bearish)} cases")
    print(f"  Bullish: {len(bullish)} cases")


if __name__ == "__main__":
    main()
